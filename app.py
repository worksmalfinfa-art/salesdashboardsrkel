"""
GROVE Sales Analytics — CLOUD Edition
======================================
Area Management Tool for GROVE at CIBIS
Cloud deployment — Supabase (PostgreSQL) + Streamlit Cloud + Google OAuth

Tech Stack: Streamlit Cloud + Supabase + Plotly + Google OAuth
Version: 2.0.0-cloud
"""

import streamlit as st
import pandas as pd
from supabase import create_client
import os
import re
import json
from datetime import datetime, timedelta, date
from io import BytesIO
import openpyxl

pd.set_option("future.no_silent_downcasting", True)

# ============================================================================
# CONFIGURATION
# ============================================================================
APP_TITLE = "GROVE Data Manager"
APP_ICON = "📊"
APP_VERSION = "2.0.0-cloud"

TIME_SEGMENTS = {
    "Breakfast": (7, 10),
    "Lunch": (12, 14),
    "After Office": (17, 19),
}

ESB_HEADER_ROW = 14
ESB_DATA_START_ROW = 15
ESB_COL_MAP = {
    "branch_name": 1, "sales_type": 2, "sales_date": 3, "sales_hour": 4,
    "bill_total": 5, "pax_total": 6, "subtotal": 7,
    "discount_total": 8, "menu_discount": 9, "nett_sales": 10,
}

COLORS = {
    "primary": "#2A31D8", "secondary": "#465FFF", "accent": "#0BA5EC",
    "gold": "#7A5AF8", "dark": "#101828", "light": "#DCEEE4",
    "red": "#F04438", "blue": "#0BA5EC", "orange": "#F79009", "white": "#FFFFFF",
}
# One palette, defined once. Charts and tenant marks draw from the same list
# so a brand keeps its hue whether it appears as a bar, a line or a square.
TENANT_HUES = ["#465FFF", "#7A5AF8", "#0BA5EC", "#12B76A", "#F79009",
               "#EE46BC", "#6172F3", "#F04438", "#36BFFA", "#9E77ED"]
CHART_PALETTE = TENANT_HUES

ALLOWED_DOMAINS = ["srkel.id", "teamup.id"]


# ============================================================================
# Supabase (PostgreSQL) DATABASE LAYER
# ----------------------------------------------------------------------------
# Public API is identical to the Google Sheets layer this replaces, so the
# parsers, dashboards, uploads and exports call it unchanged.
#
# The hand-rolled session_state cache is gone. It existed to stay under the
# Sheets quota of 60 reads/min; Postgres has no such quota, so the only thing
# still worth avoiding is refetching a whole table on every Streamlit rerun.
# st.cache_data handles that in a few lines.
# ============================================================================

# PostgREST caps every response at 1000 rows no matter what limit is requested,
# so full-table reads must page explicitly. Without this, any table past 1000
# rows would silently load only its first page.
_PAGE_SIZE = 1000
# Rows per insert request — keeps each JSON payload comfortably small.
_CHUNK_SIZE = 500


@st.cache_resource(show_spinner=False)
def _get_client():
    """One Supabase client per server process."""
    cfg = st.secrets["supabase"]
    return create_client(cfg["url"], cfg["service_key"])


@st.cache_data(ttl=300, show_spinner=False)
def _fetch_table(table, order_col=None, descending=False):
    """Read an entire table into a DataFrame, paging past the 1000-row cap."""
    client = _get_client()
    rows, start = [], 0
    while True:
        q = client.table(table).select("*")
        if order_col:
            q = q.order(order_col, desc=descending)
        page = q.range(start, start + _PAGE_SIZE - 1).execute().data or []
        rows.extend(page)
        if len(page) < _PAGE_SIZE:
            break
        start += _PAGE_SIZE
    return pd.DataFrame(rows) if rows else pd.DataFrame()


class SupabaseDB:
    _SALES_COLS = ("tenant_id", "sales_date", "sales_hour", "pax_total",
                   "subtotal", "discount_total", "nett_sales",
                   "uploaded_by", "uploaded_at")
    _PG_COLS = ("order_id", "amount", "nett_sales", "tax_amount",
                "customer_name", "sales_date", "child_total",
                "companion_total", "uploaded_by", "uploaded_at")

    def __init__(self):
        self.client = _get_client()

    # --- internals ---
    @staticmethod
    def _invalidate():
        """Drop cached reads after a write."""
        _fetch_table.clear()

    def _count(self, table):
        # head=True asks PostgREST for the count alone, with no rows in the body.
        res = self.client.table(table).select("*", count="exact", head=True).execute()
        return res.count or 0

    def _insert_chunked(self, table, records, on_conflict=None, ignore_duplicates=False):
        if not records:
            return
        for i in range(0, len(records), _CHUNK_SIZE):
            chunk = records[i:i + _CHUNK_SIZE]
            q = self.client.table(table)
            if on_conflict:
                q.upsert(chunk, on_conflict=on_conflict,
                         ignore_duplicates=ignore_duplicates).execute()
            else:
                q.insert(chunk).execute()
        self._invalidate()

    # --- User Management ---
    def get_user(self, email):
        df = _fetch_table("users")
        if df.empty: return None
        match = df[df["email"] == email]
        return match.iloc[0].to_dict() if not match.empty else None

    def create_user(self, email, display_name, role, tenant_access, created_by):
        self.client.table("users").insert({
            "email": email, "display_name": display_name, "role": role,
            "tenant_access": tenant_access, "created_by": created_by,
        }).execute()
        self._invalidate()

    def get_all_users(self):
        return _fetch_table("users", order_col="created_at")

    def has_users(self):
        return self._count("users") > 0

    def update_user_status(self, email, is_active):
        val = is_active if isinstance(is_active, bool) \
              else str(is_active).upper() in ("TRUE", "1")
        self.client.table("users").update({"is_active": val}).eq("email", email).execute()
        self._invalidate()

    def update_user_role(self, email, new_role, new_access):
        self.client.table("users").update({
            "role": new_role, "tenant_access": new_access,
        }).eq("email", email).execute()
        self._invalidate()

    # --- Tenants (the brand) ---
    def get_tenants(self, active_only=False):
        df = _fetch_table("tenants", order_col="tenant_id")
        if active_only and not df.empty and "is_active" in df.columns:
            df = df[df["is_active"] == True]
        return df

    def _next_id(self, df, col, prefix):
        """Lowest unused id, so deleting then adding cannot collide."""
        used = set(df[col].astype(str)) if not df.empty and col in df.columns else set()
        n = 1
        while f"{prefix}{n:03d}" in used:
            n += 1
        return f"{prefix}{n:03d}"

    def add_tenant(self, tenant_name, esb_branch_name=None, category="F&B", pos_type="esb"):
        tid = self._next_id(self.get_tenants(), "tenant_id", "T")
        self.client.table("tenants").insert({
            "tenant_id": tid, "tenant_name": tenant_name,
            "esb_branch_name": esb_branch_name or None,
            "category": category, "pos_type": pos_type,
        }).execute()
        self._invalidate()
        return tid

    def get_tenant_by_branch(self, esb_branch_name):
        df = self.get_tenants()
        if df.empty: return None
        match = df[df["esb_branch_name"] == esb_branch_name]
        return match.iloc[0].to_dict() if not match.empty else None

    def delete_tenant(self, tenant_id):
        res = self.client.table("sales_data").select("id", count="exact", head=True) \
                  .eq("tenant_id", tenant_id).execute()
        sales_count = res.count or 0
        out = self.client.table("tenants").delete().eq("tenant_id", tenant_id).execute()
        self._invalidate()
        return (1 if out.data else 0), sales_count

    def edit_tenant(self, tenant_id, new_name, new_branch,
                    category=None, pos_type=None, is_active=None):
        # sales_data keys on tenant_id now, so renaming a brand touches no
        # sales row at all -- the label lives in exactly one place.
        payload = {"tenant_name": new_name, "esb_branch_name": new_branch or None}
        if category is not None:  payload["category"] = category
        if pos_type is not None:  payload["pos_type"] = pos_type
        if is_active is not None: payload["is_active"] = bool(is_active)
        res = self.client.table("tenants").update(payload).eq("tenant_id", tenant_id).execute()
        self._invalidate()
        return bool(res.data)

    # --- Units (the physical space) ---
    def get_units(self):
        return _fetch_table("units", order_col="unit_id")

    def add_unit(self, unit_code, unit_name=None, floor=None, area_sqm=None):
        uid = self._next_id(self.get_units(), "unit_id", "U")
        self.client.table("units").insert({
            "unit_id": uid, "unit_code": unit_code, "unit_name": unit_name,
            "floor": floor, "area_sqm": area_sqm,
        }).execute()
        self._invalidate()
        return uid

    def edit_unit(self, unit_id, **fields):
        allowed = {k: v for k, v in fields.items()
                   if k in ("unit_code", "unit_name", "floor", "area_sqm", "is_active")}
        if not allowed: return False
        res = self.client.table("units").update(allowed).eq("unit_id", unit_id).execute()
        self._invalidate()
        return bool(res.data)

    def delete_unit(self, unit_id):
        res = self.client.table("units").delete().eq("unit_id", unit_id).execute()
        self._invalidate()
        return bool(res.data)

    # --- Tenancies (who occupied which unit, when) ---
    def get_tenancies(self):
        return _fetch_table("tenancies", order_col="start_date", descending=True)

    def get_tenancies_detailed(self):
        """Tenancies with unit code and brand name attached, for display."""
        tn = self.get_tenancies()
        if tn.empty: return tn
        out, units, tenants = tn.copy(), self.get_units(), self.get_tenants()
        if not units.empty:
            out = out.merge(units[["unit_id", "unit_code"]], on="unit_id", how="left")
        if not tenants.empty:
            out = out.merge(tenants[["tenant_id", "tenant_name"]], on="tenant_id", how="left")
        out["status"] = out["end_date"].isna().map({True: "Aktif", False: "Selesai"})
        return out

    def get_active_tenancy(self, unit_id):
        df = self.get_tenancies()
        if df.empty: return None
        m = df[(df["unit_id"] == unit_id) & (df["end_date"].isna())]
        return m.iloc[0].to_dict() if not m.empty else None

    def add_tenancy(self, unit_id, tenant_id, start_date, end_date=None, notes=None):
        self.client.table("tenancies").insert({
            "unit_id": unit_id, "tenant_id": tenant_id,
            "start_date": str(start_date),
            "end_date": str(end_date) if end_date else None,
            "notes": notes,
        }).execute()
        self._invalidate()

    def end_tenancy(self, tenancy_id, end_date):
        self.client.table("tenancies").update({"end_date": str(end_date)}) \
            .eq("tenancy_id", int(tenancy_id)).execute()
        self._invalidate()

    def replace_tenant(self, unit_id, new_tenant_id, handover_date):
        """
        Hand a unit over to a different brand.

        Closes the outgoing tenancy the day before the hand-over and opens the
        incoming one on it, so the unit is never occupied twice on the same day
        -- which the database would reject anyway. No sales row is touched:
        each brand keeps exactly what it earned, before and after.
        """
        handover = pd.to_datetime(handover_date).date()
        current = self.get_active_tenancy(unit_id)
        if current:
            started = pd.to_datetime(current["start_date"]).date()
            if started >= handover:
                raise ValueError(
                    "Tanggal serah terima harus setelah tanggal mulai penyewa "
                    f"saat ini ({started:%d %b %Y})."
                )
            self.end_tenancy(current["tenancy_id"], handover - timedelta(days=1))
        self.add_tenancy(unit_id, new_tenant_id, handover,
                         notes=f"Serah terima {handover:%Y-%m-%d}")

    # --- Targets ---
    def get_targets(self):
        return _fetch_table("targets", order_col="period_month", descending=True)

    def upsert_target(self, tenant_id, period_month, target_nett, updated_by=None):
        self.client.table("targets").upsert({
            "tenant_id": tenant_id,
            "period_month": str(period_month),
            "target_nett": int(target_nett),
            "updated_by": updated_by,
        }, on_conflict="tenant_id,period_month").execute()
        self._invalidate()

    # --- Sales Data ---
    def get_sales_data(self, tenant_filter=None):
        # Read the view, not the table: it resolves which unit each sale
        # belonged to by matching the sale's date against the tenancy periods,
        # so a hand-over splits history at the right day without the app
        # having to get that date logic right in every dashboard.
        df = _fetch_table("v_sales_enriched", order_col="sales_date")
        if df.empty: return df
        df = df.copy()
        df["sales_date"] = pd.to_datetime(df["sales_date"], errors="coerce")
        for col in ["pax_total", "subtotal", "discount_total", "nett_sales"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        if tenant_filter and tenant_filter != "All":
            df = df[df["tenant_name"] == tenant_filter]
        return df

    def append_sales_data(self, rows):
        records = [dict(zip(self._SALES_COLS, r)) for r in rows]
        # Collapse repeats of the same slot inside one file before sending.
        # Postgres refuses an upsert that touches the same key twice ("ON
        # CONFLICT DO UPDATE command cannot affect row a second time"), and an
        # ESB export can legitimately split one hour over several rows when the
        # branch records more than one sales type. Those are separate real
        # sales in the same hour, so they add up rather than overwrite.
        merged = {}
        for rec in records:
            key = (rec["tenant_id"], rec["sales_date"], rec["sales_hour"])
            if key in merged:
                for f in ("pax_total", "subtotal", "discount_total", "nett_sales"):
                    merged[key][f] += rec[f]
            else:
                merged[key] = dict(rec)
        # Upsert on the (tenant, date, hour) unique constraint. Re-uploading the
        # same file therefore rewrites those slots with identical values instead
        # of appending a second copy -- the behaviour the Sheets version lacked,
        # where every upload simply appended and a repeat doubled the day.
        self._insert_chunked("sales_data", list(merged.values()),
                             on_conflict="tenant_id,sales_date,sales_hour")

    def count_existing_slots(self, tenant_id, rows):
        """How many (date, hour) slots in this file the tenant already has.
        Lets the upload screen say what will be overwritten before it happens."""
        df = _fetch_table("sales_data")
        if df.empty or "tenant_id" not in df.columns: return 0
        have = {(str(r["sales_date"])[:10], str(r["sales_hour"]))
                for _, r in df[df["tenant_id"] == tenant_id].iterrows()}
        incoming = {(str(r["sales_date"])[:10], str(r["sales_hour"])) for r in rows}
        return len(have & incoming)

    def get_existing_keys(self, tenant_id):
        df = _fetch_table("sales_data")
        if df.empty or "tenant_id" not in df.columns: return set()
        df = df[df["tenant_id"] == tenant_id]
        return {(str(r["tenant_id"]), str(r["sales_date"])[:10], str(r["sales_hour"]))
                for _, r in df.iterrows()}

    # --- Playground Data ---
    def get_playground_data(self):
        df = _fetch_table("playground_sales", order_col="sales_date")
        if df.empty: return df
        df = df.copy()
        df["sales_date"] = pd.to_datetime(df["sales_date"], errors="coerce")
        for col in ["amount", "nett_sales", "tax_amount", "child_total", "companion_total"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
        return df

    def get_playground_tenant_id(self):
        """The tenant registered with pos_type 'playground', if any."""
        df = self.get_tenants()
        if df.empty or "pos_type" not in df.columns: return None
        m = df[df["pos_type"] == "playground"]
        return m.iloc[0]["tenant_id"] if not m.empty else None

    def insert_playground_batch(self, rows):
        records = [dict(zip(self._PG_COLS, r)) for r in rows]
        # Stamp the owning tenant so Playground is just another tenant to the
        # master dashboard rather than a hard-coded second world.
        pg_id = self.get_playground_tenant_id()
        if pg_id:
            for rec in records:
                rec["tenant_id"] = pg_id
        # order_id is the primary key, so the database rejects repeats. Collapse
        # duplicates inside the file first, since one statement cannot touch the
        # same key twice.
        seen, unique = set(), []
        for rec in records:
            oid = rec["order_id"]
            if oid in seen: continue
            seen.add(oid)
            unique.append(rec)
        before = self._count("playground_sales")
        self._insert_chunked("playground_sales", unique,
                             on_conflict="order_id", ignore_duplicates=True)
        return self._count("playground_sales") - before

    # --- Upload Log ---
    def log_upload(self, tenant_name, file_name, rows_added, uploaded_by):
        # Milliseconds included: upload_id is a primary key here, and two
        # uploads finishing in the same second would otherwise collide.
        uid = f"UPL-{datetime.now().strftime('%Y%m%d%H%M%S%f')[:-3]}"
        self.client.table("upload_log").insert({
            "upload_id": uid, "tenant_name": tenant_name, "file_name": file_name,
            "rows_added": rows_added, "uploaded_by": uploaded_by,
        }).execute()
        self._invalidate()

    def get_upload_log(self):
        return _fetch_table("upload_log", order_col="uploaded_at", descending=True)

    def get_db_stats(self):
        return {
            "sales_rows": self._count("sales_data"),
            "pg_rows": self._count("playground_sales"),
            "tenants": self._count("tenants"),
            "users": self._count("users"),
            "uploads": self._count("upload_log"),
            "db_size_kb": "supabase",
        }


# ============================================================================
# ESB FILE PARSER
# ============================================================================
class ESBParser:
    @staticmethod
    def parse(file_bytes, filename="upload.xlsx"):
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
        ws = wb.active
        meta = {
            "company": ws.cell(2, 1).value or "",
            "period": ws.cell(5, 2).value or "",
        }
        rows = []
        for r in range(ESB_DATA_START_ROW, ws.max_row + 1):
            branch = ws.cell(r, ESB_COL_MAP["branch_name"]).value
            if not branch:
                continue
            raw_date = ws.cell(r, ESB_COL_MAP["sales_date"]).value
            if isinstance(raw_date, datetime):
                sales_date = raw_date.strftime("%Y-%m-%d")
            elif isinstance(raw_date, str):
                sales_date = raw_date[:10]
            else:
                continue
            sales_hour = str(ws.cell(r, ESB_COL_MAP["sales_hour"]).value or "")
            pax = ws.cell(r, ESB_COL_MAP["pax_total"]).value or 0
            subtotal = ws.cell(r, ESB_COL_MAP["subtotal"]).value or 0
            discount = (ws.cell(r, ESB_COL_MAP["discount_total"]).value or 0) + \
                       (ws.cell(r, ESB_COL_MAP["menu_discount"]).value or 0)
            nett = ws.cell(r, ESB_COL_MAP["nett_sales"]).value or 0
            rows.append({
                "branch_name": str(branch).strip(),
                "sales_date": sales_date,
                "sales_hour": sales_hour.strip(),
                "pax_total": int(pax), "subtotal": int(subtotal),
                "discount_total": int(discount), "nett_sales": int(nett),
            })
        return rows, meta


class PlaygroundParser:
    @staticmethod
    def parse(file_bytes, filename="upload.csv"):
        import csv as csvmod
        text = file_bytes.decode("utf-8-sig")
        reader = csvmod.DictReader(text.splitlines())
        rows = []
        for r in reader:
            try:
                rows.append({
                    "order_id": r.get("Order ID","").strip(),
                    "amount": int(float(r.get("Amount",0))),
                    "nett_sales": int(float(r.get("Nett Sales",0))),
                    "tax_amount": int(float(r.get("Tax Amount",0))),
                    "customer_name": r.get("Name","").strip().strip('"'),
                    "sales_date": r.get("Date","").strip(),
                    "child_total": int(float(r.get("Child Total",0))),
                    "companion_total": int(float(r.get("Companion Total",0))),
                })
            except (ValueError, TypeError):
                continue
        return rows


def get_db():
    return SupabaseDB()


# ============================================================================
# STYLING
# ============================================================================
def apply_custom_css():
    """
    TailAdmin design tokens on Streamlit primitives.

    Structure stays Streamlit's job -- that discipline is what finally made the
    layout stable -- so nothing here sets position, height, or a negative
    margin. This block carries identity only: canvas, cards, type, badges.
    """
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@400;500;600;700;800&display=swap');
    :root{
      --bg:#F9FAFB; --card:#FFFFFF; --line:#E4E7EC;
      --ink:#101828; --muted:#667085;
      --brand:#465FFF; --brand-50:#ECF3FF;
      --up:#12B76A; --down:#F04438;
    }
    .stApp{background:var(--bg)}
    html,body,[class*="css"],button,input,select,textarea{
      font-family:'Outfit',"Segoe UI",system-ui,sans-serif!important}
    header[data-testid="stHeader"]{background:var(--bg)}

    [data-testid="stVerticalBlockBorderWrapper"]:has(> div > [data-testid="stVerticalBlock"]){
      background:var(--card);border:1px solid var(--line);border-radius:16px}

    [data-testid="stMetricValue"]{
      font-size:1.8rem;font-weight:700;color:var(--ink);
      font-variant-numeric:tabular-nums;letter-spacing:-.03em}
    [data-testid="stMetricLabel"]{font-size:.8rem;color:var(--muted);font-weight:500}
    /* Delta as a TailAdmin badge pill; direction colour stays Streamlit's. */
    [data-testid="stMetricDelta"]{
      font-size:.74rem;font-weight:600;background:var(--bg);
      border-radius:999px;padding:2px 8px;width:fit-content}

    h1,h2,h3,h4,h5,h6{color:var(--ink);letter-spacing:-.02em;font-family:'Outfit',sans-serif}
    h1{font-size:1.55rem!important;font-weight:700}
    h2{font-size:1.15rem!important;font-weight:700}
    .stCaption,[data-testid="stCaptionContainer"]{color:var(--muted)}

    section[data-testid="stSidebar"]{background:var(--card);border-right:1px solid var(--line)}
    section[data-testid="stSidebar"] [role="radiogroup"] label{
      padding:7px 12px;border-radius:10px;font-size:.86rem;color:var(--muted)}
    section[data-testid="stSidebar"] [role="radiogroup"] label:hover{background:var(--bg)}
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked){
      background:var(--brand-50);font-weight:600}
    section[data-testid="stSidebar"] [role="radiogroup"] label:has(input:checked) p{color:var(--brand)}
    section[data-testid="stSidebar"] [role="radiogroup"] input{display:none}

    .stTabs [data-baseweb="tab-list"]{gap:2px;border-bottom:1px solid var(--line)}
    .stTabs [data-baseweb="tab"]{font-size:.86rem;font-weight:500;color:var(--muted)}
    .stTabs [aria-selected="true"]{color:var(--brand)!important;font-weight:600}
    .stTabs [data-baseweb="tab-highlight"]{background-color:var(--brand)}

    /* Segmented period control (horizontal radio inside main-area cards). */
    [data-testid="stMain"] .stRadio [role="radiogroup"]{gap:4px}
    [data-testid="stMain"] .stRadio [role="radiogroup"] label{
      border:1px solid var(--line);border-radius:8px;padding:3px 12px;font-size:.8rem}
    [data-testid="stMain"] .stRadio [role="radiogroup"] label:has(input:checked){
      background:var(--brand);border-color:var(--brand)}
    [data-testid="stMain"] .stRadio [role="radiogroup"] label:has(input:checked) p{color:#fff}
    [data-testid="stMain"] .stRadio input{display:none}

    .stButton button,.stDownloadButton button{
      border-radius:10px;font-weight:600;font-size:.85rem;border:1px solid var(--line)}
    .stDataFrame{font-variant-numeric:tabular-nums}
    </style>
    """, unsafe_allow_html=True)

TENANT_HUES = ["#465FFF", "#7A5AF8", "#0BA5EC", "#12B76A", "#F79009",
               "#EE46BC", "#6172F3", "#F04438", "#36BFFA", "#9E77ED"]
CHART_PALETTE = TENANT_HUES
# Names kept so ~30 call sites need no edit; values are TailAdmin brand
# blue, dark blue, violet and grey-50.
GREEN, GREEN_D, OCHRE, FLAT = "#465FFF", "#2A31D8", "#7A5AF8", "#F2F4F7"










def card(title=None, subtitle=None):
    """
    A card is a bordered Streamlit container -- nothing else.

    Use as `with card("Title"):`. Because Streamlit builds and measures the
    container itself, its height always matches what is inside it.
    """
    c = st.container(border=True)
    if title:
        c.markdown(f"###### {title}")
    if subtitle:
        c.caption(subtitle)
    return c




DEFAULT_CHART_CONFIG = {
    "displayModeBar": True, "displaylogo": False,
    "modeBarButtonsToRemove": ["lasso2d", "select2d", "toImage",
                               "autoScale2d", "toggleSpikelines"],
    "scrollZoom": False}






# --- names the rest of the app still calls -------------------------------
def _plain(x):
    """Strip any leftover markup so old fragments render as text, not tags."""
    return re.sub(r"<[^>]+>", "", str(x)).replace("&rarr;", "->").replace("&amp;", "&").strip()











def render_header(subtitle=None):
    """No-op. The search bar and user chip were decoration built from raw HTML;
    the identity already lives in the sidebar, so the markup is simply gone."""
    return None








# ============================================================================
# AUTH PAGES
# ============================================================================
def get_user_email():
    """Try multiple methods to get the logged-in user's email."""
    try:
        if hasattr(st, "user") and hasattr(st.user, "email") and st.user.email:
            return st.user.email
    except: pass
    try:
        if hasattr(st, "user") and isinstance(st.user, dict) and st.user.get("email"):
            return st.user["email"]
    except: pass
    return st.session_state.get("manual_email", "")


def page_setup():
    """First-time setup — register the first Super Admin."""
    st.markdown('<div class="login-box"><h2>🌿 Initial Setup</h2><p>Buat akun Super Admin pertama.</p></div>', unsafe_allow_html=True)
    user_email = get_user_email()
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if user_email:
            st.info(f"Login sebagai: **{user_email}**")
        with st.form("setup_form"):
            if not user_email:
                user_email_input = st.text_input("Email Anda", placeholder="alfin@srkel.id")
            else:
                user_email_input = user_email
            display_name = st.text_input("Nama Lengkap", placeholder="e.g. Alfin")
            if st.form_submit_button("🚀 Buat Super Admin", use_container_width=True):
                email = user_email or user_email_input
                if not email or "@" not in email:
                    st.error("Email wajib diisi dengan format yang benar.")
                elif not display_name:
                    st.error("Nama lengkap wajib diisi.")
                else:
                    domain = email.split("@")[-1]
                    if domain not in ALLOWED_DOMAINS:
                        st.error(f"❌ Domain @{domain} tidak diizinkan.")
                    else:
                        get_db().create_user(email, display_name, "Super Admin", "ALL", "SYSTEM")
                        st.session_state["manual_email"] = email
                        st.success("✅ Super Admin berhasil dibuat!")
                        st.rerun()


def page_login():
    """Fallback login when Google auth doesn't provide email."""
    st.markdown('<div class="login-box"><h2>🌿 GROVE Sales Analytics</h2><p>Masukkan email terdaftar untuk melanjutkan.</p></div>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("login_form"):
            email = st.text_input("Email", placeholder="nama@srkel.id")
            if st.form_submit_button("🔐 Login", use_container_width=True):
                if not email:
                    st.error("Email wajib diisi.")
                else:
                    domain = email.split("@")[-1] if "@" in email else ""
                    if domain not in ALLOWED_DOMAINS:
                        st.error(f"❌ Domain @{domain} tidak diizinkan.")
                    else:
                        db = get_db()
                        user = db.get_user(email)
                        if not user:
                            st.error("Email tidak terdaftar. Hubungi Super Admin.")
                        elif str(user.get("is_active","TRUE")).upper() not in ("TRUE","1"):
                            st.error("Akun dinonaktifkan. Hubungi Admin.")
                        else:
                            st.session_state["manual_email"] = email
                            st.session_state["authenticated"] = True
                            st.rerun()


def check_auth():
    """Check auth — Google email or fallback manual login. Returns user dict or None."""
    user_email = get_user_email()
    if not user_email:
        return None

    domain = user_email.split("@")[-1] if "@" in user_email else ""
    if domain not in ALLOWED_DOMAINS:
        st.error(f"❌ Email {user_email} — domain @{domain} tidak diizinkan.")
        st.stop()
        return None

    db = get_db()
    user = db.get_user(user_email)
    if not user:
        if not db.has_users():
            return None
        # Auto-create Viewer for Google-auth users only
        google_email = ""
        try:
            if hasattr(st, "user") and hasattr(st.user, "email") and st.user.email:
                google_email = st.user.email
        except: pass
        if google_email:
            name = user_email.split("@")[0].replace("."," ").title()
            db.create_user(user_email, name, "Viewer", "ALL", "AUTO")
            user = db.get_user(user_email)
        else:
            return None

    if str(user.get("is_active","TRUE")).upper() not in ("TRUE","1"):
        st.error("❌ Akun Anda dinonaktifkan. Hubungi Admin.")
        st.stop()
        return None

    return user


# ============================================================================
# SIDEBAR
# ============================================================================
def render_sidebar():
    user = st.session_state["user"]
    role = user["role"]
    with st.sidebar:
        stats = get_db().get_db_stats()
        st.markdown("### GROVE")
        st.caption(f"{user['display_name']} · {role}")
        st.caption(f"F&B {stats['sales_rows']:,} · Playground {stats['pg_rows']:,}")
        st.divider()

        menu = [
            "📤 Upload F&B (ESB)",
            "📤 Upload Playground",
        ]
        if role in ("Super Admin", "Admin"):
            menu.append("🏢 Kelola Tenant")
            menu.append("🏬 Kelola Unit & Sewa")
        if role == "Super Admin":
            menu.append("👥 Kelola User")
        menu.append("📋 Upload Log")

        selected = st.radio("Menu", menu, label_visibility="collapsed")

        st.divider()
        if st.button("🚪 Logout", use_container_width=True):
            for k in list(st.session_state.keys()): del st.session_state[k]
            st.rerun()
    return selected

def page_upload_esb():
    render_header()
    db = get_db()
    user = st.session_state["user"]

    st.subheader("📤 Upload Data ESB")
    st.markdown("""
    <div class="upload-info">
        <strong>Petunjuk:</strong> Upload file Excel output dari ESB POS
        (<em>Sales Recapitulation by Time Report</em>). Sistem akan otomatis mendeteksi
        nama tenant dari kolom Branch Name dan menambahkan data baru tanpa duplikasi.<br>
        <strong>💡 Bisa upload beberapa file sekaligus!</strong>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("Pilih file ESB (.xlsx)", type=["xlsx"], accept_multiple_files=True)

    if uploaded:
        for file in uploaded:
            with st.expander(f"📄 {file.name}", expanded=True):
                try:
                    raw = file.read()
                    rows, meta = ESBParser.parse(raw, file.name)
                    if not rows:
                        st.error("File tidak berisi data valid.")
                        continue

                    branches = {}
                    for r in rows:
                        branches.setdefault(r["branch_name"], []).append(r)

                    for branch, branch_rows in branches.items():
                        tenant = db.get_tenant_by_branch(branch)
                        if not tenant:
                            st.warning(f"⚠️ Branch **{branch}** belum terdaftar.")
                            ca, cb = st.columns([3,1])
                            with ca:
                                new_name = st.text_input(f"Nama tenant untuk '{branch}':", value=branch,
                                                         key=f"n_{branch}_{file.name}")
                            with cb:
                                st.markdown("<br>", unsafe_allow_html=True)
                                if st.button("➕ Daftarkan", key=f"r_{branch}_{file.name}"):
                                    db.add_tenant(new_name, branch)
                                    st.success(f"✅ Tenant '{new_name}' terdaftar.")
                                    st.rerun()
                            continue

                        tenant_name = tenant["tenant_name"]
                        tenant_id = tenant["tenant_id"]
                        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                        insert_rows = [
                            (tenant_id, r["sales_date"], r["sales_hour"],
                             r["pax_total"], r["subtotal"], r["discount_total"],
                             r["nett_sales"], user["email"], now)
                            for r in branch_rows
                        ]

                        st.markdown(f"**Tenant:** {tenant_name} · **Data:** {len(branch_rows)} baris")

                        # Say what a repeat upload will do before it happens.
                        # The Sheets version appended blindly, so re-uploading a
                        # day doubled it; here the slots are rewritten, but the
                        # user should still see how many are affected.
                        overlap = db.count_existing_slots(tenant_id, branch_rows)
                        if overlap:
                            st.warning(
                                f"**{overlap:,} dari {len(branch_rows):,} baris** sudah ada "
                                f"di database untuk tenant ini pada tanggal dan jam yang sama. "
                                f"Baris tersebut akan **ditimpa**, bukan ditambahkan — "
                                f"jadi total penjualan tidak akan berlipat."
                            )

                        preview = pd.DataFrame(branch_rows[:5])
                        st.dataframe(preview, use_container_width=True, hide_index=True)

                        if st.button(f"✅ Upload {len(insert_rows)} baris", key=f"u_{branch}_{file.name}"):
                            with st.spinner("Menyimpan ke database..."):
                                db.append_sales_data(insert_rows)
                                added = len(insert_rows)
                                db.log_upload(tenant_name, file.name, added, user["email"])
                            st.success(f"✅ **{added}** baris berhasil diupload!")
                            if added > 0:
                                st.balloons()

                except Exception as e:
                    st.error(f"❌ Error parsing: {e}")


def page_units():
    st.markdown("## 🏬 Kelola Unit & Masa Sewa")
    st.caption(
        "Unit adalah ruang fisiknya dan sifatnya tetap; tenant adalah brand yang "
        "menempatinya dan bisa berganti. Penjualan selalu melekat pada brand, "
        "sehingga pergantian penyewa tidak pernah mengubah sejarah siapa pun."
    )
    db = get_db()
    units, tenants = db.get_units(), db.get_tenants()
    tenancies = db.get_tenancies_detailed()
    user = st.session_state["user"]

    t_occ, t_swap, t_unit, t_hist, t_target = st.tabs([
        "🏢 Okupansi", "🔄 Ganti Penyewa", "➕ Unit", "🕓 Riwayat Sewa", "🎯 Target"])

    # ---------- Okupansi saat ini ----------
    with t_occ:
        if units.empty:
            st.info("Belum ada unit. Tambahkan di tab **➕ Unit**.")
        else:
            occ = _occupancy_table(units, tenancies)
            terisi = int((occ["Status"] == "🟢 Terisi").sum())
            c1, c2, c3 = st.columns(3)
            c1.metric("Total Unit", len(occ))
            c2.metric("Terisi", terisi)
            c3.metric("Kosong", len(occ) - terisi,
                      delta=None if len(occ) == terisi else "perlu tenant")
            st.dataframe(occ, use_container_width=True, hide_index=True)

    # ---------- Ganti penyewa ----------
    with t_swap:
        if units.empty or tenants.empty:
            st.info("Perlu minimal satu unit dan satu tenant terdaftar.")
        else:
            st.markdown("#### Serah terima unit ke brand lain")
            unit_map = {f"{r['unit_code']} ({r['unit_id']})": r for _, r in units.iterrows()}
            pick = st.selectbox("Unit", list(unit_map.keys()))
            unit = unit_map[pick]
            current = db.get_active_tenancy(unit["unit_id"])

            if current:
                cur_name = tenants[tenants["tenant_id"] == current["tenant_id"]]
                cur_name = cur_name.iloc[0]["tenant_name"] if not cur_name.empty else current["tenant_id"]
                st.info(f"Penyewa saat ini: **{cur_name}** — sejak {str(current['start_date'])[:10]}")
            else:
                st.warning("Unit ini sedang kosong. Pengisian akan dicatat sebagai penyewa pertama.")

            avail = tenants if current is None else tenants[tenants["tenant_id"] != current["tenant_id"]]
            if avail.empty:
                st.error("Tidak ada brand lain yang bisa dipilih. Daftarkan dulu di **Kelola Tenant**.")
            else:
                tmap = {f"{r['tenant_name']} ({r['tenant_id']})": r for _, r in avail.iterrows()}
                new_pick = st.selectbox("Penyewa baru", list(tmap.keys()))
                new_tenant = tmap[new_pick]
                handover = st.date_input("Tanggal mulai penyewa baru", value=date.today())

                st.markdown(
                    f"Penjualan **sampai {handover - timedelta(days=1):%d %b %Y}** tetap "
                    f"tercatat atas penyewa lama, dan mulai **{handover:%d %b %Y}** "
                    f"tercatat atas **{new_tenant['tenant_name']}**. "
                    "Tidak ada satu baris penjualan pun yang diubah."
                )
                if st.button("🔄 Proses serah terima", type="primary", use_container_width=True):
                    try:
                        db.replace_tenant(unit["unit_id"], new_tenant["tenant_id"], handover)
                        st.success(
                            f"✅ **{unit['unit_code']}** kini ditempati "
                            f"**{new_tenant['tenant_name']}** sejak {handover:%d %b %Y}."
                        )
                        st.rerun()
                    except ValueError as e:
                        st.error(str(e))
                    except Exception as e:
                        st.error(f"Gagal memproses: {e}")

    # ---------- Kelola unit ----------
    with t_unit:
        st.markdown("#### Tambah unit")
        with st.form("add_unit"):
            c1, c2 = st.columns(2)
            with c1:
                u_code = st.text_input("Kode Unit", placeholder="mis. Slot DOD, A-3")
                u_floor = st.text_input("Lantai", placeholder="mis. GF")
            with c2:
                u_name = st.text_input("Keterangan", placeholder="opsional")
                u_area = st.number_input("Luas (m²)", min_value=0.0, step=0.5, value=0.0)
            if st.form_submit_button("➕ Tambah Unit", use_container_width=True):
                if not u_code:
                    st.error("Kode unit wajib diisi.")
                else:
                    db.add_unit(u_code, u_name or None, u_floor or None, u_area or None)
                    st.success(f"✅ Unit **{u_code}** ditambahkan.")
                    st.rerun()

        if not units.empty:
            st.markdown("#### Ubah unit yang ada")
            st.caption("Edit langsung di tabel, lalu tekan Simpan.")
            editable = units[["unit_id", "unit_code", "unit_name", "floor", "area_sqm", "is_active"]].copy()
            edited = st.data_editor(
                editable, hide_index=True, use_container_width=True, key="unit_editor",
                disabled=["unit_id"],
                column_config={
                    "unit_id":   st.column_config.TextColumn("ID", width="small"),
                    "unit_code": st.column_config.TextColumn("Kode Unit", required=True),
                    "unit_name": st.column_config.TextColumn("Keterangan"),
                    "floor":     st.column_config.TextColumn("Lantai", width="small"),
                    "area_sqm":  st.column_config.NumberColumn("Luas (m²)", format="%.1f"),
                    "is_active": st.column_config.CheckboxColumn("Aktif"),
                })
            if st.button("💾 Simpan perubahan unit", use_container_width=True):
                changed = 0
                for _, row in edited.iterrows():
                    before = editable[editable["unit_id"] == row["unit_id"]].iloc[0]
                    if not before.equals(row):
                        db.edit_unit(row["unit_id"], unit_code=row["unit_code"],
                                     unit_name=row["unit_name"], floor=row["floor"],
                                     area_sqm=float(row["area_sqm"]) if pd.notna(row["area_sqm"]) else None,
                                     is_active=bool(row["is_active"]))
                        changed += 1
                st.success(f"✅ {changed} unit diperbarui." if changed else "Tidak ada perubahan.")
                if changed: st.rerun()

    # ---------- Riwayat sewa ----------
    with t_hist:
        if tenancies.empty:
            st.info("Belum ada riwayat sewa.")
        else:
            show = tenancies.copy()
            show["end_date"] = show["end_date"].fillna("— sekarang")
            cols = [c for c in ["unit_code", "tenant_name", "start_date", "end_date", "status", "notes"]
                    if c in show.columns]
            st.dataframe(
                show[cols], use_container_width=True, hide_index=True,
                column_config={
                    "unit_code":   "Unit",
                    "tenant_name": "Penyewa",
                    "start_date":  "Mulai",
                    "end_date":    "Selesai",
                    "status":      "Status",
                    "notes":       "Catatan",
                })
            st.caption(
                "Baris berstatus **Aktif** adalah penyewa yang sedang menempati. "
                "Database menolak dua masa sewa yang tumpang tindih pada unit yang sama, "
                "karena tumpang tindih akan menggandakan pendapatan unit itu di setiap laporan."
            )

    # ---------- Target ----------
    with t_target:
        if tenants.empty:
            st.info("Belum ada tenant.")
        else:
            st.markdown("#### Target penjualan bulanan")
            period = st.date_input("Bulan", value=date.today().replace(day=1),
                                   help="Tanggal berapa pun; yang dipakai bulannya.")
            month_start = period.replace(day=1)

            existing = db.get_targets()
            cur = {}
            if not existing.empty:
                sel = existing[existing["period_month"].astype(str).str[:7] == f"{month_start:%Y-%m}"]
                cur = dict(zip(sel["tenant_id"], sel["target_nett"])) if not sel.empty else {}

            tgt_df = pd.DataFrame({
                "tenant_id":   tenants["tenant_id"],
                "Tenant":      tenants["tenant_name"],
                "Target (Rp)": [int(cur.get(t, 0)) for t in tenants["tenant_id"]],
            })
            edited_t = st.data_editor(
                tgt_df, hide_index=True, use_container_width=True, key="target_editor",
                disabled=["tenant_id", "Tenant"],
                column_config={
                    "tenant_id":   None,
                    "Tenant":      st.column_config.TextColumn("Tenant"),
                    "Target (Rp)": st.column_config.NumberColumn("Target (Rp)", min_value=0, step=1_000_000, format="%d"),
                })
            if st.button(f"💾 Simpan target {month_start:%B %Y}", use_container_width=True):
                n = 0
                for _, r in edited_t.iterrows():
                    db.upsert_target(r["tenant_id"], month_start, r["Target (Rp)"], user["email"])
                    n += 1
                st.success(f"✅ Target {month_start:%B %Y} tersimpan untuk {n} tenant.")
                st.rerun()


def page_tenants():
    render_header()
    db = get_db()
    st.subheader("🏢 Kelola Tenant")
    tdf = db.get_tenants()
    if not tdf.empty:
        st.dataframe(tdf, use_container_width=True, hide_index=True)

    st.divider()

    tab_add, tab_edit, tab_delete = st.tabs(["➕ Tambah Tenant", "✏️ Edit Tenant", "🗑️ Hapus Tenant"])

    # --- TAB: Tambah ---
    with tab_add:
        with st.form("add_tenant"):
            c1,c2 = st.columns(2)
            with c1: t_name = st.text_input("Nama Tenant", placeholder="e.g. Kopi Kenangan CIBIS")
            with c2: t_branch = st.text_input("ESB Branch Name (persis seperti di file)", placeholder="e.g. KOPI_KENANGAN_CIBIS")
            if st.form_submit_button("➕ Tambah Tenant", use_container_width=True):
                if t_name and t_branch:
                    db.add_tenant(t_name, t_branch)
                    st.success(f"✅ Tenant '{t_name}' berhasil ditambahkan.")
                    st.rerun()
                else:
                    st.error("Semua field wajib diisi.")

    # --- TAB: Edit ---
    with tab_edit:
        if tdf.empty:
            st.info("Belum ada tenant untuk diedit.")
        else:
            tenant_options = {f"{r['tenant_name']} ({r['tenant_id']})": r for _, r in tdf.iterrows()}
            selected = st.selectbox("Pilih Tenant", list(tenant_options.keys()), key="edit_sel")
            tenant = tenant_options[selected]

            with st.form("edit_tenant"):
                c1, c2 = st.columns(2)
                with c1:
                    new_name = st.text_input("Nama Tenant", value=tenant["tenant_name"])
                with c2:
                    new_branch = st.text_input("ESB Branch Name", value=tenant["esb_branch_name"])

                sales_df = db.get_sales_data(tenant["tenant_name"])
                sales_count = len(sales_df) if not sales_df.empty else 0
                st.caption(f"ℹ️ Tenant ini memiliki **{sales_count:,}** baris data sales. "
                           f"Jika nama diubah, seluruh data akan otomatis menyesuaikan.")

                if st.form_submit_button("💾 Simpan Perubahan", use_container_width=True):
                    if not new_name or not new_branch:
                        st.error("Semua field wajib diisi.")
                    else:
                        ok = db.edit_tenant(tenant["tenant_id"], new_name, new_branch)
                        if ok:
                            st.success(f"✅ Tenant berhasil diupdate → **{new_name}** / `{new_branch}`")
                            st.rerun()
                        else:
                            st.error("Gagal mengupdate tenant.")

    # --- TAB: Hapus ---
    with tab_delete:
        if tdf.empty:
            st.info("Belum ada tenant untuk dihapus.")
        else:
            del_options = {f"{r['tenant_name']} ({r['tenant_id']})": r for _, r in tdf.iterrows()}
            del_selected = st.selectbox("Pilih Tenant", list(del_options.keys()), key="del_sel")
            del_tenant = del_options[del_selected]

            sales_df2 = db.get_sales_data(del_tenant["tenant_name"])
            sales_count = len(sales_df2) if not sales_df2.empty else 0

            st.warning(
                f"⚠️ **Menghapus tenant '{del_tenant['tenant_name']}'** akan menghapus:\n"
                f"- Data tenant dari daftar\n"
                f"- **{sales_count:,} baris** data sales terkait\n"
                f"- Seluruh log upload terkait\n\n"
                f"**Aksi ini tidak dapat dibatalkan.**"
            )

            if st.button(f"🗑️ Ya, Hapus '{del_tenant['tenant_name']}'",
                         type="primary", use_container_width=True, key="confirm_del"):
                deleted, sales_del = db.delete_tenant(del_tenant["tenant_id"])
                if deleted:
                    st.success(f"✅ Tenant **{del_tenant['tenant_name']}** dihapus beserta {sales_del:,} baris data sales.")
                    st.rerun()
                else:
                    st.error("Gagal menghapus tenant.")


# ============================================================================
# USER MANAGEMENT
# ============================================================================
# --- Supabase Auth admin: akun login dashboard Vercel -----------------------
# Streamlit login pakai email + allowlist domain; dashboard Vercel pakai
# Supabase Auth (email + password). Kunci service_role di secrets sudah punya
# hak admin ke Auth, jadi Kelola User bisa mengurus keduanya sekaligus.
def _auth_admin():
    return _get_client().auth.admin


def auth_list_emails():
    """Set email yang sudah punya akun dashboard. Gagal -> set kosong."""
    try:
        users = _auth_admin().list_users(page=1, per_page=1000)
        return {(u.email or "").lower() for u in users if u.email}
    except Exception:
        return set()


def _auth_find(email):
    try:
        users = _auth_admin().list_users(page=1, per_page=1000)
        for u in users:
            if (u.email or "").lower() == email.lower():
                return u
    except Exception:
        pass
    return None


def auth_set_password(email, password):
    """Buat akun dashboard, atau reset password bila sudah ada."""
    u = _auth_find(email)
    if u:
        _auth_admin().update_user_by_id(u.id, {"password": password})
        return "reset"
    _auth_admin().create_user(
        {"email": email, "password": password, "email_confirm": True})
    return "baru"


def auth_set_blocked(email, blocked):
    """Blokir/buka akun dashboard. ban_duration mengikuti gotrue: '100 tahun'
    sebagai nonaktif permanen, 'none' mencabutnya."""
    u = _auth_find(email)
    if not u:
        return False
    _auth_admin().update_user_by_id(
        u.id, {"ban_duration": "876600h" if blocked else "none"})
    return True


def page_users():
    render_header()
    db = get_db()
    st.subheader("👥 Kelola User")
    st.caption(
        "Satu tempat untuk dua akses: data user aplikasi ini DAN akun login "
        "dashboard Vercel (Supabase Auth). Kolom **Dashboard** menandai siapa "
        "yang sudah bisa masuk ke dashboard."
    )
    udf = db.get_all_users()
    dash_emails = auth_list_emails()
    if not udf.empty:
        view = udf.copy()
        if "email" in view.columns:
            view["Dashboard"] = view["email"].str.lower().map(
                lambda e: "✅" if e in dash_emails else "—")
        st.dataframe(view, use_container_width=True, hide_index=True)
    st.divider()
    c1,c2 = st.columns(2)
    with c1:
        st.markdown("**Tambah / Update User**")
        st.caption("User dengan domain @srkel.id atau @teamup.id akan otomatis terdaftar sebagai Viewer saat pertama kali login. Gunakan form ini untuk mengubah role.")
        with st.form("add_user"):
            email = st.text_input("Email User", placeholder="nama@srkel.id")
            dn = st.text_input("Nama Lengkap")
            role = st.selectbox("Role", ["Viewer","Admin","Super Admin"])
            tdf = db.get_tenants()
            if not tdf.empty and role == "Viewer":
                acc = st.multiselect("Akses Tenant", tdf["tenant_name"].tolist())
                acc_str = ",".join(acc) if acc else "ALL"
            else:
                acc_str = "ALL"
                if role != "Viewer":
                    st.info("Admin & Super Admin → akses semua tenant.")
            pw = st.text_input(
                "Password dashboard Vercel (opsional)", type="password",
                help="Diisi → akun login dashboard dibuat (atau password-nya "
                     "di-reset bila sudah ada). Kosongkan → hanya data user "
                     "aplikasi ini yang disimpan. Minimal 6 karakter.")
            if st.form_submit_button("💾 Simpan", use_container_width=True):
                if email and dn:
                    existing = db.get_user(email)
                    if existing:
                        db.update_user_role(email, role, acc_str)
                        st.success(f"✅ Role '{email}' diupdate ke {role}.")
                    else:
                        db.create_user(email, dn, role, acc_str, st.session_state["user"]["email"])
                        st.success(f"✅ User '{email}' ditambahkan sebagai {role}.")
                    if pw:
                        if len(pw) < 6:
                            st.warning("⚠️ Password dashboard minimal 6 karakter — akun dashboard TIDAK dibuat.")
                        else:
                            try:
                                mode = auth_set_password(email, pw)
                                st.success("✅ Akun dashboard dibuat — bisa langsung login di Vercel."
                                           if mode == "baru" else
                                           "✅ Password dashboard di-reset.")
                            except Exception as e:
                                st.error(f"❌ Gagal membuat akun dashboard: {e}")
                    if not pw:
                        st.rerun()
    with c2:
        st.markdown("**Update Status User**")
        st.caption("Nonaktif memblokir keduanya: aplikasi ini langsung, dashboard Vercel maksimal ~1 jam (masa berlaku sesi).")
        if not udf.empty:
            with st.form("upd_user"):
                target = st.selectbox("Pilih User", udf["email"].tolist())
                new_st = st.selectbox("Status", ["TRUE","FALSE"], format_func=lambda x: "Aktif" if x=="TRUE" else "Nonaktif")
                if st.form_submit_button("🔄 Update", use_container_width=True):
                    db.update_user_status(target, new_st)
                    try:
                        if auth_set_blocked(target, new_st == "FALSE"):
                            st.success(f"✅ Status '{target}' diupdate — akun dashboard ikut "
                                       + ("diblokir." if new_st == "FALSE" else "dibuka."))
                        else:
                            st.success(f"✅ Status '{target}' diupdate (tidak punya akun dashboard).")
                    except Exception:
                        st.success(f"✅ Status '{target}' diupdate. ⚠️ Akun dashboard gagal disentuh — cek manual di Supabase.")
                    st.rerun()


# ============================================================================
# UPLOAD LOG
# ============================================================================
def page_upload_log():
    render_header()
    db = get_db()
    st.subheader("📋 Log Upload")
    log = db.get_upload_log()
    if not log.empty:
        st.dataframe(log, use_container_width=True, hide_index=True)
    else:
        st.info("Belum ada riwayat upload.")


# ============================================================================
# UPLOAD PLAYGROUND
# ============================================================================
def page_upload_playground():
    render_header()
    db = get_db()
    user = st.session_state["user"]

    st.subheader("🎪 Upload Data Playground TnT")
    st.markdown("""
    <div class="upload-info">
        <strong>Petunjuk:</strong> Upload file CSV dari POS Playground
        (<em>pos_sales_report</em>). Sistem akan otomatis mendeteksi
        dan menambahkan data baru berdasarkan Order ID (tanpa duplikasi).<br>
        <strong>💡 Bisa upload beberapa file sekaligus!</strong>
    </div>
    """, unsafe_allow_html=True)

    uploaded = st.file_uploader("Pilih file Playground (.csv)", type=["csv"], accept_multiple_files=True, key="pg_upload")

    if uploaded:
        for file in uploaded:
            with st.expander(f"📄 {file.name}", expanded=True):
                try:
                    raw = file.read()
                    rows = PlaygroundParser.parse(raw, file.name)
                    if not rows:
                        st.error("File tidak berisi data valid.")
                        continue

                    st.markdown(f"**Data ditemukan:** {len(rows)} transaksi")
                    date_min = min(r["sales_date"] for r in rows)
                    date_max = max(r["sales_date"] for r in rows)
                    total_amount = sum(r["amount"] for r in rows)
                    total_child = sum(r["child_total"] for r in rows)
                    st.markdown(f"**Periode:** {date_min} s.d. {date_max} · **Total:** Rp {total_amount:,} · **Children:** {total_child:,}")

                    preview = pd.DataFrame(rows[:5])
                    st.dataframe(preview, use_container_width=True, hide_index=True)

                    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    insert_rows = [(r["order_id"],r["amount"],r["nett_sales"],r["tax_amount"],
                                    r["customer_name"],r["sales_date"],r["child_total"],
                                    r["companion_total"],user["email"],now) for r in rows]

                    if st.button(f"✅ Upload {len(insert_rows)} transaksi", key=f"pg_{file.name}"):
                        with st.spinner("Menyimpan ke database..."):
                            added = db.insert_playground_batch(insert_rows)
                            db.log_upload("Playground TnT", file.name, added, user["email"])
                        skipped = len(insert_rows) - added
                        st.success(f"✅ **{added}** transaksi baru" + (f" · {skipped} duplikat di-skip" if skipped else ""))
                        if added > 0: st.balloons()
                except Exception as e:
                    st.error(f"❌ Error parsing: {e}")


def main():
    st.set_page_config(page_title=APP_TITLE, page_icon=APP_ICON, layout="wide", initial_sidebar_state="expanded")
    apply_custom_css()

    db = get_db()

    # First contact with the database. Failing here almost always means the
    # secrets are wrong or schema.sql was never run, so say that plainly
    # instead of surfacing a raw PostgREST error.
    try:
        has_users = db.has_users()
    except Exception as e:
        st.error(
            "**Tidak bisa terhubung ke database.**\n\n"
            "Periksa dua hal:\n\n"
            "1. `url` dan `service_key` di bagian `[supabase]` pada Streamlit Secrets "
            "sudah benar — pastikan memakai kunci **service_role**, bukan `anon`.\n"
            "2. `supabase/schema.sql` sudah dijalankan di Supabase Studio → SQL Editor."
        )
        st.caption(f"Detail teknis — {type(e).__name__}: {e}")
        return

    # First-time setup
    if not has_users:
        page_setup()
        return

    # Auth check
    user = check_auth()
    if not user:
        page_login()
        return

    st.session_state["user"] = user

    sel = render_sidebar()
    if "Upload F&B" in sel: page_upload_esb()
    elif "Upload Playground" in sel: page_upload_playground()
    elif "Unit" in sel: page_units()
    elif "Tenant" in sel: page_tenants()
    elif "User" in sel: page_users()
    elif "Log" in sel: page_upload_log()


if __name__ == "__main__":
    main()
